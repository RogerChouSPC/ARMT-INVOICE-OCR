"""
Test script: process each sample invoice with improved Gemini extraction.
Reads digital PDFs with pdfplumber. Scanned PDFs are noted as needing Vision.
Reports results in table format. Does NOT deploy anything.
"""
import sys, json, os, re, base64, time, urllib.request, urllib.error
sys.stdout.reconfigure(encoding='utf-8')

import pdfplumber
import openpyxl
from datetime import datetime

GEMINI_API_KEY = "AIzaSyAv_Dvb8CS1oEdzGpcjnP1Txl6P3Fc_qVU"
GEMINI_MODEL   = "gemini-2.5-flash"
INVOICE_DIR    = "ARMT-INVOICE-OCR/data/sample_invoices"
OUR_TAXID      = "0107537001421"  # สหพัฒนพิบูล

# ── Load Customer Master ──────────────────────────────────────────────────────
def load_customer_master():
    wb = openpyxl.load_workbook("ARMT-INVOICE-OCR/data/Customer Master.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v for v in row):
            r = dict(zip(headers, row))
            rows.append({
                "store_name":    (r.get("ชื่อร้านค้า") or "").strip(),
                "customergroup": r.get("customergroup") or "",
                "customercode":  r.get("customercode") or "",
                "taxid":         r.get("taxid") or "",
            })
    return rows

CUSTOMER_MASTER = load_customer_master()
CUSTOMER_MASTER_JSON = json.dumps(CUSTOMER_MASTER, ensure_ascii=False)

# ── Build system prompt ───────────────────────────────────────────────────────
SYSTEM_PROMPT = f"""You are an expert at extracting structured data from Thai invoice text (ใบแจ้งหนี้, ใบวางบิล, ใบเสร็จรับเงิน).

CRITICAL CONTEXT — READ CAREFULLY:
- บริษัท สหพัฒนพิบูล จำกัด (มหาชน) (taxid 0107537001421) is OUR COMPANY — the RECIPIENT of invoices.
- The company that ISSUED / SENT the invoice to us is the VENDOR/CUSTOMER in output fields.
- NEVER put สหพัฒนพิบูล in customergroup, customercode, or taxid output fields.
- taxid in output = the invoice ISSUER's tax ID (NOT 0107537001421).

CUSTOMER MASTER LOOKUP TABLE (use this to fill customergroup and customercode):
{CUSTOMER_MASTER_JSON}

HOW TO MAP customergroup and customercode:
1. Find the invoice issuer's taxid (the 13-digit number that is NOT 0107537001421).
2. Find the issuer's company name and branch from the invoice header.
3. Look up the CUSTOMER MASTER table above: match by taxid first, then refine by company name/branch.
   - If the same taxid has multiple entries (e.g., ซีพี แอ็กซ์ตร้า), pick the row whose store_name best matches the issuer's name in the invoice.
   - For example: taxid 0107567000414 with "Lotus" or "โลตัส" or "LT" → group "05 - โลตัส"
   - taxid 0107567000414 with "Makro" or "แม็คโคร" → group "04 - ซีพี แอ็กซ์ตร้า(Makro)"
   - taxid 0107536000633 with "คลังครอสด็อคธัญบุรี 00485" → customercode "0115526"
   - taxid 0107536000633 headquarters → customercode "0102856"
   - taxid 0105540016253 (City Mall / ซิตี้มอลล์) with branch 00001 → customercode "0114682"
4. Copy customergroup and customercode EXACTLY from the matching Customer Master row.

HOW TO MAP vendor_customercode:
- Look for: "รหัสร้านค้า", "Customer Code", "ชื่อลูกค้า [CODE]", "เจ้าของ/ตัวแทน(รหัสร้านค้า)"
- Strip any text prefixes: "BTM-MC15009" → "15009"; "TOP-M802316" → "802316"; "CFW-M900548" → "900548"; "LS16985" → "LS16985" (keep if no clear rule); numeric only is preferred.
- For Big C: the code starting with "4000047" or similar 7-digit number next to our name.
- For CP All / 7-11: the supplier code (2000087 etc.) from the invoice.

HOW TO MAP vendor_branch:
- Look for: "Group [number]", "สาขาที่ [code]", "Branch", "Site code", or branch number in company header.
- Use the branch code WITHOUT leading zeros where it's a number (e.g., "00485" stays "00485", "29130" stays "29130").
- If vendor is สำนักงานใหญ่ (head office) with no branch code: use "00000".

Given raw text from one or more invoice pages (separated by "--- PAGE BREAK ---"), extract ALL invoice line items and return a JSON array. Each element = one row.

Output fields (22 columns):
- customergroup: from Customer Master lookup (exact copy)
- customercode: from Customer Master lookup (exact copy)
- taxid: the invoice ISSUER's 13-digit tax ID
- vendor_customercode: our code in vendor's system (strip text prefixes, keep numbers)
- vendor_branch: vendor branch/group code for our transactions
- vendor_expensecode: expense code if present, else ""
- vendor_expensegroup: expense group if present, else ""
- divisionsale: division or sale code if present, else ""
- invoiceno: invoice number / เลขที่
- invoicedate: invoice date → YYYY-MM-DD (Buddhist Era: subtract 543 from year)
- duedate: due/payment date → YYYY-MM-DD
- description: main invoice description / purpose line
- product_description: product or service line item detail
- amount: amount before deductions (plain number, no commas)
- vat_7: VAT 7% amount (number or "0")
- tax_pct: general tax % amount (number or "0")
- tax_2: withholding tax 2% (number or "0")
- tax_3: withholding tax 3% (number or "0")
- tax_5: withholding tax 5% (number or "0")
- netamount: net payable after all deductions (number)
- remark: any notes / หมายเหตุ

Rules:
- Multiple line items per invoice → one row per item (share header: invoiceno, taxid, dates, etc.)
- Missing fields → empty string ""
- Numbers → plain string without commas or currency symbols
- Dates → YYYY-MM-DD; Buddhist Era year → subtract 543
- Tax ID = exactly 13 digits
- Return ONLY a valid JSON array, no markdown fences, no explanation"""


CALL_DELAY = 8  # seconds between Gemini calls (10 RPM limit → 6s min, use 8s for safety)

def gemini_post(payload: dict, timeout: int = 90, retries: int = 3) -> dict:
    """Raw Gemini API call with retry on 429 (waits 65s before retry)."""
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent?key={GEMINI_API_KEY}"
    for attempt in range(retries):
        req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers={"Content-Type": "application/json"})
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return json.loads(resp.read())
        except urllib.error.HTTPError as e:
            body = e.read().decode()
            if e.code == 429 and attempt < retries - 1:
                print(f"    [429 rate limit, waiting 65s before retry {attempt+2}/{retries}]")
                time.sleep(65)
                continue
            raise
    raise RuntimeError("Max retries exceeded")


def call_gemini_text(text: str, filename: str) -> list:
    """Send invoice text to Gemini for field extraction."""
    truncated = text[:12000] + "\n[truncated]" if len(text) > 12000 else text
    payload = {
        "system_instruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": [{"role": "user", "parts": [{"text": f"Filename: {filename}\n\nInvoice text:\n\n{truncated}"}]}],
        "generationConfig": {
            "temperature": 0.1,
            "responseMimeType": "application/json",
            "thinkingConfig": {"thinkingBudget": 0},
        },
    }
    try:
        data = gemini_post(payload, timeout=60)
        raw = data["candidates"][0]["content"]["parts"][0]["text"]
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, list) else [parsed]
    except Exception as ex:
        return [{"error": str(ex)}]
    finally:
        time.sleep(CALL_DELAY)


def call_gemini_vision(base64_img: str, filename: str) -> str:
    """OCR a single page image with Gemini Vision."""
    payload = {
        "contents": [{"parts": [
            {"inline_data": {"mime_type": "image/png", "data": base64_img}},
            {"text": "Extract ALL text from this invoice image exactly as it appears. Include every word, number, date, and special character. Preserve structure and line breaks. Output only the raw extracted text."},
        ]}],
        "generationConfig": {"temperature": 0, "thinkingConfig": {"thinkingBudget": 0}},
    }
    try:
        data = gemini_post(payload, timeout=90)
        return data["candidates"][0]["content"]["parts"][0]["text"]
    except Exception as ex:
        return f"[OCR ERROR: {ex}]"
    finally:
        time.sleep(CALL_DELAY)


def extract_text(path: str) -> tuple[str, bool, int]:
    """Extract text from PDF. Returns (text, is_digital, page_count)."""
    with pdfplumber.open(path) as pdf:
        parts = []
        total_chars = 0
        for page in pdf.pages:
            t = page.extract_text() or ""
            parts.append(t)
            total_chars += len(t.replace(" ", "").replace("\n", ""))
        avg = total_chars / max(len(pdf.pages), 1)
        is_digital = avg > 100
        return "\n\n--- PAGE BREAK ---\n\n".join(parts), is_digital, len(pdf.pages)


def render_page_base64(path: str, page_idx: int) -> str:
    """Render a PDF page to base64 PNG using pdfplumber + PIL."""
    try:
        import io
        with pdfplumber.open(path) as pdf:
            page = pdf.pages[page_idx]
            img = page.to_image(resolution=150)
            buf = io.BytesIO()
            img.save(buf, format="PNG")
            return base64.b64encode(buf.getvalue()).decode()
    except Exception as e:
        return ""


def print_table(rows: list, filename: str):
    cols = [
        ("customergroup", 40), ("customercode", 45), ("taxid", 15),
        ("vendor_customercode", 22), ("vendor_branch", 14),
        ("invoiceno", 25), ("invoicedate", 12), ("duedate", 12),
        ("amount", 12), ("vat_7", 8), ("tax_3", 8), ("netamount", 12),
        ("description", 50),
    ]
    print(f"\n{'='*130}")
    print(f"FILE: {filename}")
    print(f"{'='*130}")
    if not rows:
        print("  (no rows)")
        return
    if "error" in rows[0]:
        print(f"  ERROR: {rows[0]['error']}")
        return
    # Header
    header = " | ".join(f"{c[:w]:<{w}}" for c, w in cols)
    print(header)
    print("-" * len(header))
    for r in rows:
        row_str = " | ".join(f"{str(r.get(c,''))[:w]:<{w}}" for c, w in cols)
        print(row_str)


# ── Main loop ─────────────────────────────────────────────────────────────────
files = sorted(os.listdir(INVOICE_DIR))
results = {}

for fname in files:
    if not fname.endswith(".pdf"):
        continue
    path = os.path.join(INVOICE_DIR, fname)
    print(f"\nProcessing {fname}...")

    text, is_digital, pages = extract_text(path)

    if is_digital:
        print(f"  → Digital PDF ({pages} pages, text extracted)")
        rows = call_gemini_text(text, fname)
    else:
        print(f"  → Scanned PDF ({pages} pages), running Vision OCR...")
        ocr_texts = []
        for p in range(pages):
            b64 = render_page_base64(path, p)
            if not b64:
                ocr_texts.append("")
                continue
            ocr_text = call_gemini_vision(b64, fname)
            ocr_texts.append(ocr_text)
            print(f"    page {p+1}/{pages} OCR done")
            time.sleep(3)
        combined = "\n\n--- PAGE BREAK ---\n\n".join(ocr_texts)
        print(f"  → Extracting fields from OCR text...")
        rows = call_gemini_text(combined, fname)

    results[fname] = rows
    print_table(rows, fname)

print("\n\n" + "="*130)
print("SUMMARY")
print("="*130)
for fname, rows in results.items():
    if rows and "error" not in rows[0]:
        cg = rows[0].get("customergroup", "?")
        cc_short = str(rows[0].get("customercode", "?"))[:40]
        print(f"  {fname:<35} → {len(rows)} rows | {cg} | {cc_short}")
    elif rows and "error" in rows[0]:
        print(f"  {fname:<35} → ERROR: {rows[0]['error']}")
    else:
        print(f"  {fname:<35} → (empty)")
